
import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side

def procesar_a_excel(input_file, output_excel):
    try:
        # Leer el archivo con el delimitador correcto (;) o '¬' dependiendo si es 20232 0 20231 respectivamente
    
        #df = pd.read_csv(input_file, sep=';', encoding='utf-8')
        df = pd.read_csv(input_file, sep='¬', encoding='utf-8', engine='python')
        
        # Limpieza básica de datos
        df = df.apply(lambda x: x.str.strip() if x.dtype == 'object' else x)
        df.replace(['', 'NA', 'N/A', 'NaN'], np.nan, inplace=True)
        
        # Crear archivo Excel con openpyxl para mejor control
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos ICFES"
        
        # Estilos para los encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                          right=Side(style='thin'), 
                          top=Side(style='thin'), 
                          bottom=Side(style='thin'))
        
        # Escribir los datos
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Encabezados
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
        
        # Ajustar el ancho de las columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar encabezados
        ws.freeze_panes = 'A2'
        
        # Guardar el archivo
        os.makedirs(os.path.dirname(output_excel), exist_ok=True)
        wb.save(output_excel)
        
        print("\n" + "="*50)
        print("PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        print(f"Archivo Excel generado en: {output_excel}")
        print("="*50 + "\n")
        
    except Exception as e:
        print("\n" + "="*50)
        print("ERROR EN EL PROCESAMIENTO")
        print(f"Error: {str(e)}")
        print("="*50 + "\n")
        raise

def main():
    input_file = 'tu_data.txt'
    output_excel = '/output/icfes_data.xlsx'
    
    procesar_a_excel(input_file, output_excel)

if __name__ == "__main__":
    main()